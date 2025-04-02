import openpyxl
import os
from copy import copy
from datetime import datetime, timedelta
from chinese_calendar import is_holiday, is_workday
import logging
from logging.handlers import RotatingFileHandler


class InspectionSheetUpdater:
    def __init__(self, target_dir='./', file_prefix="日常点检表_厚板L2"):
        """
        初始化点检表更新器

        Args:
            target_dir (str): 文件存储目录
            file_prefix (str): 文件前缀(例如"日常点检表_厚板L2")
        """
        self.target_dir = target_dir
        self.file_prefix = file_prefix
        self._setup_logging()

    def _setup_logging(self):
        """配置日志系统"""
        log_file = os.path.join(self.target_dir, "file_collection.log")
        logging.basicConfig(
            level=logging.INFO,
            format="%(asctime)s [%(levelname)s] %(message)s",
            handlers=[
                RotatingFileHandler(
                    log_file,
                    maxBytes=10 * 1024 * 1024,
                    backupCount=5
                ),
                logging.StreamHandler()
            ]
        )

    @staticmethod
    def is_special_holiday(date):
        """检查是否是特殊节假日"""
        special_holidays = [
            (1, 1),  # 元旦
            (5, 1),  # 劳动节
            (10, 1),  # 国庆节
        ]
        return (date.month, date.day) in special_holidays

    def is_workday_including_custom(self, date):
        """判断是否是工作日(包括自定义节假日)"""
        if self.is_special_holiday(date):
            return False
        return is_workday(date)

    def get_current_month_filepath(self, date=None):
        """获取当前月份文件的完整路径"""
        if date is None:
            date = datetime.now().date()
        return os.path.join(
            self.target_dir,
            f"{self.file_prefix}({date.year}-{date.month}月份).xlsx"
        )

    def create_new_month_file(self, last_month_file, new_month_file, last_day):
        """创建新月份文件并复制上月最后Sheet"""
        try:
            wb_last = openpyxl.load_workbook(last_month_file)
            sheet_names = [int(s) for s in wb_last.sheetnames if s.isdigit()]

            if not sheet_names:
                logging.error("上月文件中没有数字命名的工作表！")
                return False

            max_sheet_name = max(sheet_names)
            last_day_sheet = wb_last[str(max_sheet_name)]

            # 创建新工作簿
            wb_new = openpyxl.Workbook()
            del wb_new[wb_new.sheetnames[0]]

            # 复制工作表
            today = datetime.now().date()
            today_day = str(today.day)
            new_sheet = wb_new.create_sheet(title=today_day)

            # 复制单元格内容、样式和格式
            for row in last_day_sheet.iter_rows():
                for cell in row:
                    new_cell = new_sheet.cell(row=cell.row, column=cell.column)
                    new_cell.value = cell.value
                    if cell.has_style:
                        new_cell.font = copy(cell.font)
                        new_cell.border = copy(cell.border)
                        new_cell.fill = copy(cell.fill)
                        new_cell.number_format = cell.number_format
                        new_cell.protection = copy(cell.protection)
                        new_cell.alignment = copy(cell.alignment)

            # 复制行高和列宽
            for idx, row_dim in last_day_sheet.row_dimensions.items():
                new_sheet.row_dimensions[idx] = copy(row_dim)
            for idx, col_dim in last_day_sheet.column_dimensions.items():
                new_sheet.column_dimensions[idx] = copy(col_dim)

            # 复制合并单元格
            for merged_cell in last_day_sheet.merged_cells.ranges:
                new_sheet.merge_cells(str(merged_cell))

            # 复制打印设置
            new_sheet.page_setup = copy(last_day_sheet.page_setup)
            new_sheet.print_options = copy(last_day_sheet.print_options)
            new_sheet.print_title_rows = last_day_sheet.print_title_rows
            new_sheet.print_title_cols = last_day_sheet.print_title_cols

            wb_new.save(new_month_file)
            logging.info(f"成功创建新月份文件 {new_month_file}，并复制 {last_day} 号工作表为 {today_day} 号")
            return True

        except Exception as e:
            logging.error(f"创建新月份文件时出错: {e}")
            return False

    @staticmethod
    def auto_save_excel(filepath):
        """使用 Excel 自动打开并保存文件（修复文件状态）"""
        try:
            from win32com.client import Dispatch
            excel = Dispatch("Excel.Application")
            excel.Visible = False
            wb = excel.Workbooks.Open(os.path.abspath(filepath))
            wb.Save()
            wb.Close()
            excel.Quit()
            return True
        except Exception as e:
            logging.error(f"自动保存失败（可能无Excel环境）: {e}")
            return False

    def update_sheet(self):
        """主功能：更新点检表"""
        today = datetime.now().date()

        if not self.is_workday_including_custom(today):
            logging.info(f"{today} 是节假日或周末，不需要更新点检表")
            return False

        current_month_file = self.get_current_month_filepath()
        today_day = today.day

        # 判断是否需要创建新月份文件
        if not os.path.exists(current_month_file):
            last_month = (today.replace(day=1) - timedelta(days=1))
            last_month_file = self.get_current_month_filepath(last_month)

            if not os.path.exists(last_month_file):
                logging.error(f"上月文件 {last_month_file} 不存在！")
                return False

            return self.create_new_month_file(
                last_month_file,
                current_month_file,
                last_month.day
            )

        # 处理已有月份文件
        try:
            # 尝试自动修复文件状态
            if not self.auto_save_excel(current_month_file):
                logging.warning(f"警告：无法自动修复 {current_month_file}，尝试直接读取")

            wb = openpyxl.load_workbook(current_month_file)
            sheet_names = [int(s) for s in wb.sheetnames if s.isdigit()]

            if not sheet_names:
                logging.error(f"{current_month_file} 中没有数字命名的工作表！")
                return False

            if today_day in sheet_names:
                logging.warning(f"今天 {today_day} 号的sheet已存在，不需要更新")
                return False

            # 找到前一天的工作表（可能跳过节假日）
            previous_day = today_day - 1
            while previous_day > 0 and previous_day not in sheet_names:
                previous_day -= 1

            if previous_day < 1:
                logging.error("找不到前一天的工作表作为模板！")
                return False

            # 复制工作表
            yesterday_sheet = wb[str(previous_day)]
            new_sheet = wb.copy_worksheet(yesterday_sheet)
            new_sheet.title = str(today_day)

            wb.save(current_month_file)
            logging.info(f"成功：已复制 {previous_day} 号工作表为 {today_day} 号")
            return True

        except Exception as e:
            logging.error(f"处理文件时出错: {e}")
            return False
def main():
    updater = InspectionSheetUpdater(
        target_dir='./',
        file_prefix="日常点检表_厚板L2"
    )
    # 执行更新
    updater.update_sheet()
if __name__ == "__main__":
    main()
